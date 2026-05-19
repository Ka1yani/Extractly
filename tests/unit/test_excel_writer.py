"""Tests for the ExcelWriter service."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from extractly.models.template import MasterTemplate, Section, Field
from extractly.models.extraction import ExtractedDocument, ExtractedSection
from extractly.services.excel_writer import ExcelWriter


@pytest.fixture
def template() -> MasterTemplate:
    return MasterTemplate(
        document_type="Certificate",
        sections=[
            Section(name="Header", fields=[
                Field(label="Batch No"), Field(label="Date", field_type="date"),
            ]),
            Section(name="Results", fields=[Field(label="Assay", field_type="number")]),
        ],
    )


@pytest.fixture
def documents() -> list[ExtractedDocument]:
    return [
        ExtractedDocument(source_filename="doc1.pdf", sections=[
            ExtractedSection(name="Header", fields={"Batch No": "B-001", "Date": "2024-01-15"}),
            ExtractedSection(name="Results", fields={"Assay": "99.2%"}),
        ]),
        ExtractedDocument(source_filename="doc2.pdf", sections=[
            ExtractedSection(name="Header", fields={"Batch No": "B-002", "Date": "2024-01-16"}),
            ExtractedSection(name="Results", fields={"Assay": "98.8%"}),
        ]),
    ]


@pytest.fixture
def writer() -> ExcelWriter:
    return ExcelWriter()


class TestExcelWriter:
    def test_creates_file(self, writer, documents, template, tmp_output_dir):
        output = tmp_output_dir / "test.xlsx"
        assert writer.write(documents, template, output).exists()

    def test_header_row(self, writer, documents, template, tmp_output_dir):
        output = tmp_output_dir / "test.xlsx"
        writer.write(documents, template, output)
        wb = load_workbook(str(output))
        ws = wb.active
        assert ws is not None
        headers = [c.value for c in ws[1]]
        assert headers == ["Source File", "Header — Batch No", "Header — Date", "Results — Assay"]

    def test_data_rows(self, writer, documents, template, tmp_output_dir):
        output = tmp_output_dir / "test.xlsx"
        writer.write(documents, template, output)
        wb = load_workbook(str(output))
        ws = wb.active
        assert ws is not None
        assert [c.value for c in ws[2]] == ["doc1.pdf", "B-001", "2024-01-15", "99.2%"]
        assert [c.value for c in ws[3]] == ["doc2.pdf", "B-002", "2024-01-16", "98.8%"]

    def test_empty_documents(self, writer, template, tmp_output_dir):
        output = tmp_output_dir / "test.xlsx"
        writer.write([], template, output)
        wb = load_workbook(str(output))
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 1

    def test_missing_field_value(self, writer, template, tmp_output_dir):
        docs = [ExtractedDocument(source_filename="p.pdf", sections=[
            ExtractedSection(name="Header", fields={"Batch No": "B-003"}),
        ])]
        output = tmp_output_dir / "test.xlsx"
        writer.write(docs, template, output)
        wb = load_workbook(str(output))
        ws = wb.active
        assert ws is not None
        assert [c.value for c in ws[2]] == ["p.pdf", "B-003", None, None]

    def test_bold_headers(self, writer, documents, template, tmp_output_dir):
        output = tmp_output_dir / "test.xlsx"
        writer.write(documents, template, output)
        wb = load_workbook(str(output))
        ws = wb.active
        assert ws is not None
        for cell in ws[1]:
            assert cell.font.bold is True
