"""Task 3A: Master Spreadsheet Generator.

Aggregates extracted data from all documents into a single Excel file.
Each row = one document, each column = one field from the template.

Usage:
    writer = ExcelWriter()
    writer.write(documents, template, Path("output/master_output.xlsx"))
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from extractly.models.template import MasterTemplate
from extractly.models.extraction import ExtractedDocument

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Writes extracted data to an Excel spreadsheet.

    Flattens the hierarchical section/field structure into a flat table
    where columns are "Section — Field Label" and rows are documents.
    """

    def write(
        self,
        documents: list[ExtractedDocument],
        template: MasterTemplate,
        output_path: Path,
    ) -> Path:
        """Write all extracted data to an Excel file.

        Args:
            documents: List of extracted documents to include.
            template: The MasterTemplate (defines column headers).
            output_path: Path for the output .xlsx file.

        Returns:
            The path where the file was saved.
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        assert ws is not None, "Workbook should always have a default active sheet"
        ws.title = "Extracted Data"

        # Build headers: Source File + all field labels from template
        headers = ["Source File"] + template.all_field_labels()
        ws.append(headers)

        # Style header row (bold)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Build section→field mapping for lookup
        section_fields = self._build_section_fields(template)

        # Add one row per document
        for doc in documents:
            row = [doc.source_filename]
            for section_name, field_label in section_fields:
                value = doc.get_field_value(section_name, field_label)
                row.append(value)
            ws.append(row)

        # Auto-adjust column widths (approximate)
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(15, len(header) + 2)

        wb.save(str(output_path))
        logger.info(
            "Excel file saved: %s (%d documents, %d columns)",
            output_path,
            len(documents),
            len(headers),
        )
        return output_path

    def _build_section_fields(
        self, template: MasterTemplate
    ) -> list[tuple[str, str]]:
        """Build an ordered list of (section_name, field_label) pairs.

        Args:
            template: The MasterTemplate to flatten.

        Returns:
            List of (section_name, field_label) tuples.
        """
        pairs: list[tuple[str, str]] = []
        for section in template.sections:
            for field in section.fields:
                pairs.append((section.name, field.label))
        return pairs
